# wap to read columns from an excel and create a VCF out of columns

import openpyxl
from openpyxl import load_workbook

def createVCF(row):
    fd_vcf = open("contacts.vcf",'w')
    row[0].value

def readExcel(filename1):
    wb = load_workbook(filename = filename1)
    sheet_ranges = wb[wb.get_sheet_names()[0]]

    for row in sheet_ranges.iter_rows():
        createVCF(row)

    '''sheet = wb.active
    cells = sheet['A1':'D5']
    for c1, c2, c3, c4 in cells:
        print("{0:8} {1:8} {2:8} {3:8}".format(c1.value, c2.value, c3.value, c4.value))
'''
def main():
    filename1 = input("Enter filename to import : ")
    readExcel(filename1)

if __name__ == "__main__":
    main()